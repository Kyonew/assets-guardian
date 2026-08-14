import datetime
import hashlib
import json
import logging
import math
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell, WriteOnlyCell
from openpyxl.formatting.rule import Rule
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from assets_guardian.core.config.loader import load_profiles
from assets_guardian.core.domain.ports.sheet_builders import ISheetBuilder

from .reader import read_workbook
from .rules_loader import load_rules

logger = logging.getLogger(__name__)

# Sheet protection is UI-only (blocks editing in Excel), not real security:
# openpyxl's protection is trivially bypassable and never verified by this tool
# itself, so the password never needs to be known or communicated to anyone.
_PROTECTION_PASSWORD = "placeholder"  # noqa: S105  # nosec B105

# Name of the custom document property storing the integrity checksum of the
# auto-generated sheets (everything except "... Matrix" sheets, which are
# meant to be edited by hand and are therefore excluded from the checksum).
_CHECKSUM_PROPERTY_NAME = "assets_guardian_checksum"


class ExcelWorksheet:
    """Wrapper around an openpyxl sheet to simplify writing and tracking."""

    _LINE_HEIGHT = 15  # points per line (Calibri 11pt)

    def __init__(self, sheet: Any):
        self.sheet = sheet
        self.row_count = 0
        self.column_indices: dict[str, int] = {}  # Column name -> 1-based index
        self.auto_row_height: bool = False
        self._column_widths: list[int | float] = []

    @property
    def title(self) -> str:
        """Name of the sheet."""
        return str(self.sheet.title)

    def set_column_widths(self, widths: list[int] | list[float]) -> None:
        """Sets the column widths by index (1-based)."""
        self._column_widths = list(widths)
        for i, width in enumerate(widths, 1):
            letter = get_column_letter(i)
            self.sheet.column_dimensions[letter].width = width

    def append_row(self, values: list[Any], is_header: bool = False) -> None:
        """Appends a styled row as a WriteOnlyCell."""
        cells = []
        for i, val in enumerate(values, 1):
            cell = WriteOnlyCell(self.sheet, value=val)
            if is_header:
                self.__apply_header_style(cell)
                if isinstance(val, str):
                    self.column_indices[val] = i
            else:
                self.__apply_cell_style(cell, val)
            cells.append(cell)

        if self.auto_row_height and not is_header:
            height = self._estimate_row_height(values)
            self.sheet.row_dimensions[self.row_count + 1].height = height
        self.sheet.append(cells)
        self.row_count += 1

    def _estimate_row_height(self, values: list[Any]) -> float:
        """Estimates the row height needed to display all cell values with wrap_text."""
        max_lines = 1
        for i, val in enumerate(values):
            if not isinstance(val, str) or not val:
                continue
            col_width = self._column_widths[i] if i < len(self._column_widths) else 20
            chars_per_line = max(1, col_width)
            lines = sum(
                max(1, math.ceil(len(segment) / chars_per_line)) for segment in val.split("\n")
            )
            max_lines = max(max_lines, lines)
        return max_lines * self._LINE_HEIGHT

    @staticmethod
    def __apply_header_style(cell: Cell) -> None:
        """Applies the standard header style to cells."""
        cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    @staticmethod
    def __apply_cell_style(cell: Cell, value: Any) -> None:
        """Applies the standard style and formatting to data cells."""

        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Formatting based on type
        if isinstance(value, datetime.datetime):
            cell.number_format = "dd/mm/yyyy hh:mm:ss"
        elif isinstance(value, datetime.date):
            cell.number_format = "dd/mm/yyyy"
        elif isinstance(value, bool):
            cell.value = 1 if value else 0
            cell.number_format = "0"
        elif isinstance(value, (int, float, str)):
            formats = {int: "0", float: "0.00", str: "@"}
            cell.number_format = formats.get(type(value), "@")


class ExcelWriter:
    """Implementation of the IWriter port for Excel."""

    def __init__(self, sheet_builders: list[ISheetBuilder] | None = None):
        """Initializes the writer with a list of sheet builders."""
        self.sheet_builders = sheet_builders or []

    def write(
        self,
        data: Any,
        output_path: str | Path,
        existing_file_path: str | Path | None = None,
        rules_file_path: str | Path | None = None,
        employees_file_path: str | None = None,
    ) -> None:
        """Writes a complete Excel workbook by delegating tasks to builders."""

        self.employees_file_path = employees_file_path
        logger.info("Starting workbook writing: %s", output_path)

        rules = self.__load_all_rules(rules_file_path)
        existing_data = self.__load_existing_data(existing_file_path)
        self.__verify_integrity(existing_file_path, existing_data)

        workbook: Workbook = openpyxl.Workbook(write_only=True)
        for builder in self.sheet_builders:
            self.__process_builder(workbook, builder, data, existing_data, rules)

        workbook.save(output_path)
        self.__finalize_integrity_signature(output_path)
        size = Path(output_path).stat().st_size
        logger.info("File created: %s (size: %.2f MB)", output_path, size / 1024 / 1024)

    def __load_all_rules(self, rules_file_path: str | Path | None) -> dict[str, Any]:
        """Loads global rules and merges them with plugin rules."""

        rules = load_rules(rules_file_path) if rules_file_path else {}
        for builder in self.sheet_builders:
            try:
                rules.update(builder.get_rules())
            except Exception:
                logger.exception("Plugin rules error: %s", builder.__class__.__name__)
        return rules

    def __load_existing_data(self, path: str | Path | None) -> dict[str, Any]:
        """Loads data from an existing Excel file."""

        if path and Path(path).exists():
            try:
                return read_workbook(path)
            except Exception:
                logger.exception("Error reading existing file: %s", path)
        return {}

    def __verify_integrity(self, path: str | Path | None, existing_data: dict[str, Any]) -> None:
        """Warns if any sheet was altered outside this tool.

        Compares, per sheet, the checksum stored in the existing file (from
        the previous write) against a freshly computed one over its current
        content. A mismatch on a given sheet means it was edited (or the
        file replaced) between the last write and now — including "...
        Matrix" sheets, which are otherwise left unprotected and editable on
        purpose: this check still reports on them, it just never blocks the
        edit itself.

        Args:
            path: Path to the existing Excel file, if any.
            existing_data: Content of the existing file, as returned by
                `read_workbook` (empty dict if there was no existing file).
        """
        if not path or not existing_data:
            return

        stored_checksums, last_modified_by, modified_at = self.__read_integrity_metadata(path)
        if not stored_checksums:
            return

        altered_sheets = [
            name
            for name, sheet in existing_data.items()
            if name in stored_checksums
            and self.__compute_checksum(self.__flatten_sheet(sheet)) != stored_checksums[name]
        ]

        if altered_sheets:
            logger.warning(
                "Integrity check failed for %s: the following sheets were modified "
                "outside Assets Guardian since the last 'sync' (file last saved by "
                "%s on %s): %s.",
                path,
                last_modified_by or "an unknown user",
                modified_at or "an unknown date",
                ", ".join(altered_sheets),
            )

    @staticmethod
    def __flatten_sheet(sheet: dict[str, Any]) -> list[list[Any]]:
        """Rebuilds the [header_row, *content_rows] shape from `read_workbook` output."""
        sorted_indices = sorted(sheet["header"].keys())
        header_row = [sheet["header"][i]["title"] for i in sorted_indices]
        return [header_row, *sheet["content"]]

    def __read_integrity_metadata(
        self, path: str | Path
    ) -> tuple[dict[str, str], str | None, datetime.datetime | None]:
        """Reads the stored per-sheet checksums and last-save metadata of an existing file.

        The author/date come from the file's standard OOXML core properties
        (`lastModifiedBy`, `modified`), filled in automatically by Excel on
        every save. This is self-reported, not cryptographically verified -
        useful context for a human reading the warning, not proof of anything.
        """
        try:
            workbook = openpyxl.load_workbook(path, read_only=True)
            try:
                checksums: dict[str, str] = {}
                for prop in workbook.custom_doc_props.props:  # type: ignore[attr-defined]
                    if prop.name == _CHECKSUM_PROPERTY_NAME:
                        checksums = dict(json.loads(str(prop.value)))
                        break
                return checksums, workbook.properties.lastModifiedBy, workbook.properties.modified
            finally:
                workbook.close()
        except Exception:
            logger.exception("Error reading integrity metadata from %s", path)
        return {}, None, None

    @staticmethod
    def __compute_checksum(rows: list[list[Any]]) -> str:
        """Computes a deterministic checksum over a single sheet's rows."""
        canonical = json.dumps(rows, sort_keys=True, default=str)
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def __finalize_integrity_signature(self, output_path: str | Path) -> None:
        """Stamps the just-saved file with a per-sheet checksum of its content.

        Re-reads the file that was just saved (rather than the in-memory
        values passed to `append_row`) so the checksum reflects exactly what
        ends up on disk, e.g. booleans are stored as 0/1 in cells but were
        still plain Python bools when written — computing the checksum from
        the raw written values would then never match what a later `read_workbook`
        sees, causing permanent false-positive integrity warnings.
        """
        saved_data = read_workbook(output_path)
        checksums = {
            name: self.__compute_checksum(self.__flatten_sheet(sheet))
            for name, sheet in saved_data.items()
        }

        workbook = openpyxl.load_workbook(output_path)
        workbook.custom_doc_props.append(  # type: ignore[attr-defined]
            StringProperty(name=_CHECKSUM_PROPERTY_NAME, value=json.dumps(checksums))
        )
        workbook.save(output_path)

    def __process_builder(
        self,
        workbook: Workbook,
        builder: ISheetBuilder,
        data: Any,
        existing_data: dict[str, Any],
        rules: dict[str, Any],
    ) -> None:
        """Processes all sheets for a builder."""

        sheet_names = list(builder.sheet_names)
        if builder.source_name and builder.source_name != "default":
            instance_id = getattr(builder, "instance_id", "")
            if instance_id:
                matrix_name = f"{builder.source_name.capitalize()} ({instance_id}) Matrix"
            else:
                matrix_name = f"{builder.source_name.capitalize()} Matrix"
            if matrix_name not in sheet_names:
                sheet_names.append(matrix_name)

        for name in sheet_names:
            if name.endswith(" Matrix"):
                existing_key = (
                    name
                    if name in existing_data
                    else (name.lower() if name.lower() in existing_data else None)
                )
                if existing_key:
                    logger.info("Preserving matrix: %s", existing_key)
                    self.__copy_sheet(workbook, name, existing_data[existing_key])
                    continue

            self.__build_sheet(workbook, builder, name, data, existing_data, rules)

    def __build_sheet(
        self,
        workbook: Workbook,
        builder: ISheetBuilder,
        sheet_name: str,
        data: Any,
        existing_data: dict[str, Any],
        rules: dict[str, Any],
    ) -> None:
        """Builds a specific sheet."""

        logger.info("Processing sheet: %s", sheet_name)
        preserved = self.__get_preserved_for_sheet(existing_data, builder, sheet_name)
        sheet_rules = rules.get(sheet_name, {})

        sheet = workbook.create_sheet(sheet_name)
        worksheet = ExcelWorksheet(sheet)
        builder.build(worksheet, data, preserved, sheet_rules)

        if worksheet.row_count == 0 and sheet_name.lower().endswith(" matrix"):
            self.__initialize_matrix_sheet(worksheet, builder.source_name)

        self.__apply_rules(worksheet, sheet_rules)

        if not sheet_name.lower().endswith(" matrix"):
            self.__protect_sheet(worksheet)

    def __protect_sheet(self, worksheet: ExcelWorksheet) -> None:
        """Locks a sheet against edits within Excel's UI.

        This is a deterrent, not real security: openpyxl's protection is
        trivially bypassable and this tool never needs to unlock it itself,
        since the whole workbook is rebuilt from scratch on every write.
        """
        worksheet.sheet.protection.sheet = True
        worksheet.sheet.protection.set_password(_PROTECTION_PASSWORD)

    def __copy_sheet(self, workbook: Workbook, sheet_name: str, sheet_info: dict[str, Any]) -> None:
        """Copies an existing sheet as-is into the new workbook."""
        sheet = workbook.create_sheet(sheet_name)
        worksheet = ExcelWorksheet(sheet)

        header_info = sheet_info.get("header", {})
        if not header_info:
            return

        # Sort columns by index for the header
        sorted_indices = sorted(header_info.keys())
        header_titles = [header_info[i]["title"] for i in sorted_indices]

        # If it is a matrix, apply original widths (70 for profiles, 40 for scopes)
        if sheet_name.lower().endswith(" matrix"):
            widths = [70] + [40] * (len(sorted_indices) - 1)
        else:
            widths = [header_info[i].get("width", 15) for i in sorted_indices]

        # Write header and configure widths
        worksheet.append_row(header_titles, is_header=True)
        worksheet.set_column_widths(widths)

        # Write content rows
        for row in sheet_info.get("content", []):
            worksheet.append_row(row)

    def __get_preserved_for_sheet(
        self, existing_data: dict[str, Any], builder: ISheetBuilder, sheet_name: str
    ) -> dict[tuple[Any, ...], dict[str, Any]]:
        """Extracts the data to be preserved for a given sheet."""

        if sheet_name not in existing_data:
            return {}

        preserved_config = builder.preserved_columns.get(sheet_name, {})
        primary_keys = preserved_config.get("primary_keys", [])
        if not primary_keys:
            return {}

        sheet_info = existing_data[sheet_name]
        # Index by primary key to retrieve the row data even if its position changes
        return self.__index_content_by_primary_keys(
            sheet_info["content"], sheet_info["header"], primary_keys
        )

    def __index_content_by_primary_keys(
        self,
        content: list[list[Any]],
        header: dict[int, dict[str, Any]],
        primary_keys: list[str],
    ) -> dict[tuple[Any, ...], dict[str, Any]]:
        """Indexes rows to find them by their unique identifier.

        Structure: { (Key_Value,): { "Column": "Full_Value", ... } }
        Keeps the entire row to retrieve any cell value later.
        """

        # Map column names to their indices
        column_to_index = {}
        for index, info in header.items():
            title = info["title"]
            column_to_index[title] = index - 1

        # Retrieve the indices of primary keys
        key_indices = []
        for key in primary_keys:
            if key in column_to_index:
                key_indices.append(column_to_index[key])

        if not key_indices:
            return {}

        indexed_mapping = {}
        for row in content:
            # row -> [val1, val2, ...]
            extraction = self.__extract_row_with_key(row, key_indices, header)
            if extraction:
                # composite_key: (pk1, pk2), row_data: {"Col": val}
                composite_key, row_data = extraction
                indexed_mapping[composite_key] = row_data
        return indexed_mapping

    def __extract_row_with_key(
        self, row: list[Any], key_indices: list[int], header: dict[int, dict[str, Any]]
    ) -> tuple[tuple[Any, ...], dict[str, Any]] | None:
        """Prepares the identifier (key) and data dictionary for a row.

        E.g. row=["ID01", "Srv1", "Note"] ->
            ( ("ID01",), {"ID": "ID01", "Name": "Srv1", "Note": "Note"} )
        The tuple (key) is used for indexing, the dict contains all content.
        """
        try:
            # key_values: tuple of PK column values (e.g., ("ID-001",))
            key_values = tuple(row[i] if i < len(row) else None for i in key_indices)
            if self.__is_key_valid(key_values):
                # Reconstruct a {Title: Value} dictionary for this row
                # index - 1 to convert from 1-based (header) to 0-based (row list)
                row_data = {
                    info["title"]: (row[index - 1] if index - 1 < len(row) else None)
                    for index, info in header.items()
                }
                return key_values, row_data
        except (IndexError, TypeError):
            pass
        return None

    def __is_key_valid(self, composite_key: tuple[Any, ...]) -> bool:
        """Checks if the composite key is valid (not empty)."""
        return any(value not in (None, "") for value in composite_key)

    def __apply_rules(self, worksheet: ExcelWorksheet, rules: Any) -> None:
        """Applies validation rules and conditional formatting to a sheet."""
        if not isinstance(rules, list) or worksheet.row_count <= 1:
            return

        for col_def in rules:
            self.__process_column_rules(worksheet, col_def, rules)

    def __process_column_rules(
        self, worksheet: ExcelWorksheet, col_def: Any, all_column_rules: list[Any]
    ) -> None:
        """Processes the rule(s) of a column."""

        if not isinstance(col_def, dict) or "column_name" not in col_def:
            return

        # Extract the rules list (either via 'rules' key or the object itself)
        rules_list = col_def.get("rules")
        if not isinstance(rules_list, list):
            rules_list = [col_def] if "rule_type" in col_def else []

        for rule in rules_list:
            if isinstance(rule, dict):
                # Inject the column name so the rule knows where to apply
                rule.setdefault("column_name", col_def["column_name"])
                self.__apply_single_rule(worksheet, rule, all_column_rules)

    def __get_column_index(
        self, column_name: str, worksheet: ExcelWorksheet, all_rules: list[Any]
    ) -> int | None:
        """Determines the 1-based index of a column."""

        # Priority to detected header
        idx = worksheet.column_indices.get(column_name)
        if idx is not None:
            return idx

        # Fallback to JSON position
        for i, col_def in enumerate(all_rules):
            if isinstance(col_def, dict) and col_def.get("column_name") == column_name:
                return i + 1
        return None

    def __apply_single_rule(
        self, worksheet: ExcelWorksheet, rule: dict[str, Any], all_column_rules: list[Any]
    ) -> None:
        """Dispatches rule application depending on its type."""

        col_index = self.__get_column_index(
            rule.get("column_name", ""), worksheet, all_column_rules
        )
        if col_index is None:
            return

        col_letter = get_column_letter(col_index)
        cells_range = f"{col_letter}2:{col_letter}{worksheet.row_count}"
        rule_type = rule.get("rule_type")

        if rule_type == "list_validation":
            self.__apply_data_validation(worksheet, rule, cells_range)
        elif rule_type == "conditional_format":
            self.__apply_conditional_format(worksheet, rule, cells_range, col_letter)

    def __apply_data_validation(
        self, worksheet: ExcelWorksheet, rule: dict[str, Any], cells_range: str
    ) -> None:
        """Applies a dropdown list validation."""

        dv = DataValidation(
            type=rule.get("validate", "list"),
            formula1=f'"{",".join(rule.get("source", []))}"',
            allow_blank=rule.get("ignore_blank", True),
        )
        if hasattr(worksheet.sheet, "add_data_validation"):
            worksheet.sheet.add_data_validation(dv)
        else:
            worksheet.sheet.data_validations.append(dv)
        dv.add(cells_range)

    def __format_formula_item(self, item: Any) -> str:
        """Formats an individual item to be injected into an Excel formula.

        This method supports basic Python types and converts them to the
        textual format expected by Excel formula syntax:
        - Booleans: converted to "TRUE" or "FALSE" (uppercase).
        - Numbers (int, float): converted to strings without quotes.
        - Others (str, etc.): surrounded by double quotes.

        Args:
            item: The Python item to format.

        Returns:
            str: The formatted textual representation of the item for Excel.
        """
        if isinstance(item, bool):
            return "TRUE" if item else "FALSE"
        if isinstance(item, (int, float)):
            return str(item)
        return f'"{item}"'

    def __build_list_formula(
        self, criteria: str, formatted_items: list[str], col_letter: str
    ) -> str:
        """Builds an Excel OR or AND expression formula from a list of items.

        - For an equality criterion (==), generates an "OR" formula checking if the cell
          equals at least one item from the list.
        - For an inequality criterion (!=), generates an "AND" formula checking if the cell
          differs from all items in the list.

        Args:
            criteria: The comparison operator ("==" or "!=").
            formatted_items: The list of items already formatted for Excel formulas.
            col_letter: The letter representing the targeted column (e.g. "A").

        Returns:
            str: The final Excel expression formula (e.g. 'AND(A2<>"Val1", A2<>"Val2")').
        """
        if criteria == "==":
            cond_list = [f"{col_letter}2={item}" for item in formatted_items]
            return f"OR({', '.join(cond_list)})" if len(cond_list) > 1 else cond_list[0]

        cond_list = [f"{col_letter}2<>{item}" for item in formatted_items]
        return f"AND({', '.join(cond_list)})" if len(cond_list) > 1 else cond_list[0]

    def __build_cf_rule(
        self, criteria: str | None, val: Any, col_letter: str, style: DifferentialStyle
    ) -> Rule:
        """Generates the appropriate openpyxl conditional formatting rule (Rule).

        Based on the provided comparison criteria, this method instantiates either a rule
        based on a custom expression (type="expression") or a standard cell value comparison
        rule (type="cellIs").

        Args:
            criteria: The comparison criteria ("is_empty", "is_not_empty", "==", "!=").
            val: The comparison value or list of values.
            col_letter: The letter representing the targeted column (e.g. "A").
            style: The differential style to apply (DifferentialStyle).

        Returns:
            Rule: The Rule object configured for openpyxl conditional formatting.
        """
        if criteria == "is_empty":
            return Rule(type="expression", formula=[f"LEN(TRIM({col_letter}2))=0"], dxf=style)

        if criteria == "is_not_empty" or not criteria:
            return Rule(type="expression", formula=[f"LEN(TRIM({col_letter}2))>0"], dxf=style)

        if isinstance(val, list):
            formatted_items = [self.__format_formula_item(item) for item in val]
            formula = self.__build_list_formula(criteria or "", formatted_items, col_letter)
            return Rule(type="expression", formula=[formula], dxf=style)

        operator: Literal["equal", "notEqual"] = "equal" if criteria == "==" else "notEqual"
        if isinstance(val, str):
            val = f'"{val}"'
        return Rule(type="cellIs", operator=operator, formula=[str(val)], dxf=style)

    def __apply_conditional_format(
        self, worksheet: ExcelWorksheet, rule: dict[str, Any], cells_range: str, col_letter: str
    ) -> None:
        """Applies a complete conditional formatting rule to a range of cells.

        This method extracts the differential style based on the defined color name,
        generates the corresponding formatting rule (Rule), and registers it on the
        specified cell range of the Excel sheet.

        Args:
            worksheet: The Excel worksheet currently being edited.
            rule: The dictionary describing the conditional rule (criteria, values, formats).
            cells_range: The cell range targeted by the rule (e.g., "A2:A10").
            col_letter: The letter representing the targeted column (e.g., "A").
        """
        color_name = rule.get("format") or rule.get("color")
        style = self.__get_differential_style(color_name) if color_name else None
        if not style:
            return

        cf_rule = self.__build_cf_rule(
            rule.get("criteria"), rule.get("value", ""), col_letter, style
        )
        worksheet.sheet.conditional_formatting.add(cells_range, cf_rule)

    def __get_differential_style(self, color_name: str) -> DifferentialStyle | None:
        """Returns a differential style based on the color name (standard palette)."""
        palette = {
            "green": {"bg": "C6EFCE", "font": "006100"},
            "yellow": {"bg": "FFEB9C", "font": "9C5700"},
            "red": {"bg": "FFC7CE", "font": "9C0006"},
            "orange": {"bg": "FBE2D5", "font": "BD5015"},
            "black": {"bg": "000000", "font": "FFFFFF"},
            "blue": {"bg": "CFEEFC", "font": "145F82"},
            "white": {"bg": "FFFFFF", "font": "000000"},
        }

        if color_name not in palette:
            return None

        colors = palette[color_name]
        return DifferentialStyle(
            fill=PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid"),
            font=Font(color=colors["font"], bold=True),
        )

    def __initialize_matrix_sheet(
        self,
        worksheet: ExcelWorksheet,
        source_name: str,
    ) -> None:
        """Initializes a matrix sheet."""

        # Prepare example headers (Scopes)
        header = [
            f"{source_name.capitalize()} Matrix",
            "Group: Example1",
            "Group: Example2",
            "Instance",
            "Project: Example",
        ]

        # Widths: 70 for the first column, 40 for others
        widths: list[int] | list[float] = [70] + [40] * (len(header) - 1)
        worksheet.set_column_widths(widths)
        worksheet.append_row(header, is_header=True)

        # Prepare rows (Profiles)
        # Retrieve profiles from employees.json
        default_profiles = load_profiles(self.employees_file_path)

        for profile in default_profiles:
            worksheet.append_row([profile])

        logger.info(
            "Pivot matrix initialized for %s with profiles from employees.json.", source_name
        )
