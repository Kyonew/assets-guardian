import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


def __extract_clean_path(app_config: Any) -> str | None:
    """Extracts the cleaned path from the application configuration if present.

    Args:
        app_config: Application configuration.

    Returns:
        str | None: The cleaned Excel file path, or None if not configured.
    """
    val = app_config.paths.excel.clean_path
    if val is None:
        return None
    return str(val)


def __resolve_default_excel_path() -> Path:
    """Resolves the default Excel file path from the application configuration.

    Returns:
        Path: The resolved Excel file path.
    """
    default_path = Path("outputs/assets_guardian.xlsx")
    config_path = Path("config/config.yml")
    if not config_path.exists():
        return default_path

    try:
        from assets_guardian.core.config.app_config import AppConfig
        from assets_guardian.core.config.loader import load_yaml_config

        raw_data = load_yaml_config(str(config_path))
        app_config = AppConfig.create_from_dict(raw_data)
        clean_path = __extract_clean_path(app_config)
        if clean_path:
            return Path(clean_path)
    except Exception:
        logger.exception("Error resolving default Excel path")
    return default_path


def __normalize_sheet_names(sheet_names: list[str] | None) -> set[str] | None:
    """Normalizes the requested sheet names for robust comparison.

    Args:
        sheet_names: Optional list of sheets to normalize.

    Returns:
        set[str] | None: Set of lowercase normalized sheet names, or None.
    """
    if sheet_names is None:
        return None
    return {name.lower() for name in sheet_names}


def __should_skip_sheet(name: str, target_names: set[str] | None) -> bool:
    """Determines if a sheet should be skipped based on specified targets.

    Args:
        name: Name of the sheet to test.
        target_names: Set of normalized target sheet names.

    Returns:
        bool: True if the sheet should be skipped, False otherwise.
    """
    if target_names is None:
        return False
    return name.lower() not in target_names


def __is_non_empty_row(row: Any) -> bool:
    """Checks if a row is not empty and contains at least one non-null value.

    Args:
        row: Row of cells to analyze.

    Returns:
        bool: True if the row contains at least one non-empty cell, False otherwise.
    """
    if not row:
        return False
    return any(cell.value is not None for cell in row)


def __find_first_non_empty_row(rows_iter: Any) -> Any:
    """Finds and returns the first non-empty row from an iterator.

    Args:
        rows_iter: Iterator over the sheet's rows.

    Returns:
        Any: The first non-empty row found, or None if all rows are empty.
    """
    for row in rows_iter:
        if __is_non_empty_row(row):
            return row
    return None


def __parse_header(header_row: Any) -> dict[int, dict[str, Any]]:
    """Parses the header row to generate the header structure dictionary.

    Args:
        header_row: Row of cells representing the header.

    Returns:
        dict[int, dict[str, Any]]: Dictionary structuring the header (index -> info).
    """
    header: dict[int, dict[str, Any]] = {}
    if not header_row:
        return header

    for col_idx, cell in enumerate(header_row, 1):
        title = cell.value
        if title:
            header[col_idx] = {"title": str(title), "width": 15}
    return header


def __extract_row_values(row: Any, header_len: int) -> list[Any]:
    """Extracts values from a row based on the header length.

    Args:
        row: Row of cells to extract.
        header_len: Header length for column framing.

    Returns:
        list[Any]: List of extracted values, padded with None if necessary.
    """
    line = []
    for col_idx in range(1, header_len + 1):
        idx = col_idx - 1
        value = row[idx].value if idx < len(row) else None
        line.append(value)
    return line


def __parse_content(rows_iter: Any, header_len: int) -> list[list[Any]]:
    """Continuously reads and extracts subsequent content rows.

    Args:
        rows_iter: Iterator over the sheet's rows.
        header_len: Expected header length.

    Returns:
        list[list[Any]]: List of extracted content rows.
    """
    content = []
    for row in rows_iter:
        if __is_non_empty_row(row):
            line = __extract_row_values(row, header_len)
            content.append(line)
    return content


def __read_sheet(sheet: Any) -> dict[str, Any]:
    """Reads a specific sheet and returns its header and content.

    Args:
        sheet: openpyxl worksheet to read.

    Returns:
        dict[str, Any]: Dictionary containing the sheet's header and content.
    """
    rows_iter = iter(sheet.iter_rows(values_only=False))
    header_row = __find_first_non_empty_row(rows_iter)
    header = __parse_header(header_row)
    content = __parse_content(rows_iter, len(header))
    return {"header": header, "content": content}


def read_workbook(
    file_path: str | Path | None = None,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    """Reads an existing Excel workbook in a highly optimized manner.

    This function utilizes openpyxl's `read_only` mode to stream only target
    sheets (if provided), minimizing RAM and CPU usage.

    Args:
        file_path: Path to the Excel file. If None, resolves path from config.yml or fallback.
        sheet_names: Optional list of sheets to load. If None, all sheets are loaded.

    Raises:
        FileNotFoundError: If the specified or resolved Excel file is not found.

    Returns:
        dict[str, Any]: Workbook content structured as
            { sheet_name: {"header": ..., "content": ...} }
    """
    path = __resolve_default_excel_path() if file_path is None else Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    target_names = __normalize_sheet_names(sheet_names)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    workbook_content = {}

    try:
        for name in wb.sheetnames:
            if __should_skip_sheet(name, target_names):
                continue

            workbook_content[name] = __read_sheet(wb[name])
    finally:
        wb.close()

    return workbook_content
